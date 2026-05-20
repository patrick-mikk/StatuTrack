#!/usr/bin/env python3
"""
laws_xml_to_excel.py
====================
Convert a Department of Justice "Laws of Canada" XML statute or regulation
(e.g. https://laws-lois.justice.gc.ca/eng/XML/P-24.501.xml) into a structured
Excel workbook for compliance use.

The workbook contains:
    - README          : how the workbook is organized and what each sheet holds.
    - Metadata        : statute-level identifiers, dates, source URL.
    - Structure       : Parts, Divisions, and other headings as an outline.
    - Provisions      : every section/subsection/paragraph/subparagraph/clause
                        as a row, with a built citation (e.g. "s. 5(h)(i)") and
                        a hyperlink back to the official text.
    - Sections        : one row per Section with marginal note and dates,
                        useful as a quick index.
    - Definitions     : every defined term found in the Act, with the section
                        where it is defined and the full text of the definition.
    - Amendments      : every HistoricalNote item broken out, sortable by
                        amendment date, with citations.

Designed to be re-run periodically. The output filename includes the
consolidation date so successive runs can be diffed to track changes.

Author: prepared for Patrick (Compliance Associate, Alterna)
Usage:
    python laws_xml_to_excel.py
    python laws_xml_to_excel.py --url https://laws-lois.justice.gc.ca/eng/XML/B-1.01.xml
    python laws_xml_to_excel.py --file ./P-24.501.xml --output pcmltfa.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    from lxml import etree
except ImportError:
    sys.exit("ERROR: lxml is required. Install with: pip install lxml")

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    sys.exit("ERROR: openpyxl is required. Install with: pip install openpyxl")


LIMS_NS = "http://justice.gc.ca/lims"
DEFAULT_URL = "https://laws-lois.justice.gc.ca/eng/XML/P-24.501.xml"

# Justice Laws pages follow predictable URL patterns. Acts and regulations use
# different paths under /eng/, so build_section_url() dispatches on document type.
ACT_SECTION_URL = "https://laws-lois.justice.gc.ca/eng/acts/{code}/section-{label}.html"
REGULATION_SECTION_URL = "https://laws-lois.justice.gc.ca/eng/regulations/{code}/section-{label}.html"


# ----------------------------------------------------------------------------
# Helpers for working with the LIMS-flavoured XML
# ----------------------------------------------------------------------------

def lims(attr: str) -> str:
    """Return a Clark-notation attribute name for the LIMS namespace."""
    return f"{{{LIMS_NS}}}{attr}"


def lims_attr(elem, attr: str, default: str = "") -> str:
    """Pull a `lims:foo` attribute off an element, returning default if absent."""
    if elem is None:
        return default
    return elem.get(lims(attr), default)


def inner_text(elem) -> str:
    """Concatenate all text content inside an element (including inline tags
    such as XRefExternal, DefinedTermEn, Repealed, DefinitionRef, Language).

    The Justice Laws XML uses these inline tags inside <Text> for visual styling
    and cross-references; for our flat database we just want the readable string.

    Regulations and order-in-council preambles include <Footnote> children
    inline within <Text> — these contain the footnote definition (citation
    text), not part of the readable provision content, so we skip them.
    Similarly, <FootnoteRef> wraps the visible footnote marker letter; that
    marker isn't useful in a flat extract and is dropped.
    """
    if elem is None:
        return ""
    parts: list[str] = []
    if elem.text:
        parts.append(elem.text)
    for child in elem:
        if child.tag in ("Footnote", "FootnoteRef"):
            # Skip the footnote body and any inline marker, but keep whatever
            # text appears AFTER the child in the parent's tail.
            if child.tail:
                parts.append(child.tail)
            continue
        parts.append(inner_text(child))
        if child.tail:
            parts.append(child.tail)
    raw = "".join(parts)
    return re.sub(r"\s+", " ", raw).strip()


def own_text(elem) -> str:
    """Get the immediate <Text> child of a structural element, fully flattened."""
    if elem is None:
        return ""
    text_child = elem.find("Text")
    return inner_text(text_child)


def collect_full_text(elem) -> str:
    """Build the full text for a provision row — the element's own <Text>
    plus the labelled text of every nested Subsection/Paragraph/Subparagraph/
    Clause descendant in document order. Each child is prefixed with its
    label so structure is preserved.

    Used in the Provisions sheet's Text column so a Section row contains the
    full enumerated sub-content, not just the leading clause. The cell wraps
    so Excel renders the embedded newlines as visible line breaks.

    Example output for a Section with one subsection (1) containing
    paragraphs (a) and (b):

        Opening clause of the section.
        (1) Opening clause of subsection 1.
        (a) Text of paragraph (a).
        (b) Text of paragraph (b).

    Definition and HistoricalNote children are intentionally excluded —
    they have their own sheets.
    """
    if elem is None:
        return ""
    parts: list[str] = []

    own = own_text(elem)
    if own:
        parts.append(own)

    for child in elem:
        if child.tag in ("Subsection", "Paragraph", "Subparagraph", "Clause"):
            label = inner_text(child.find("Label"))
            child_full = collect_full_text(child)
            if not child_full:
                continue
            parts.append(f"{label} {child_full}" if label else child_full)
        elif child.tag == "ContinuedText":
            # Trailing text after enumerated paragraphs ("...all of the above").
            ct = own_text(child)
            if ct:
                parts.append(ct)

    return "\n".join(parts)


# ----------------------------------------------------------------------------
# Data model
# ----------------------------------------------------------------------------

@dataclass
class ProvisionRow:
    """One row in the Provisions sheet — represents any level of the hierarchy."""
    citation: str
    section_url: str
    level: str               # Section, Subsection, Paragraph, Subparagraph, Clause, Definition
    part: str = ""
    division: str = ""
    subdivision: str = ""
    section: str = ""
    subsection: str = ""
    paragraph: str = ""
    subparagraph: str = ""
    clause: str = ""
    marginal_note: str = ""
    text: str = ""
    in_force_date: str = ""
    last_amended_date: str = ""
    enacted_date: str = ""
    fid: str = ""
    repealed: bool = False


@dataclass
class AmendmentRow:
    citation: str
    section_label: str
    in_force_date: str
    enacted_date: str
    enact_id: str
    text: str
    citation_chunks: list[str] = field(default_factory=list)  # split on ';' per LIMS2HTML.xsl


@dataclass
class DefinitionRow:
    term_en: str
    term_fr: str
    citation: str
    section_url: str
    text: str
    in_force_date: str
    enacted_date: str
    repealed: bool


@dataclass
class HeadingRow:
    level: int
    label: str
    title: str
    in_force_date: str
    enacted_date: str


@dataclass
class EnablingAuthorityRow:
    """One row in the EnablingAuthorities sheet — for regulations only."""
    chapter_code: str = ""   # e.g. "B-1.01"
    title: str = ""           # e.g. "BANK ACT"
    reference_type: str = ""  # e.g. "act" or "regulation"


@dataclass
class ParsedDocument:
    """Top-level parsed representation of either a Statute or a Regulation."""
    # --- Polymorphic identity (set by parse_xml) ---
    document_type: str = "statute"   # "statute" | "regulation"
    url_code: str = ""               # used by build_section_url and as filename stem
                                     # "P-24.501" for an act, "SOR-2021-181" for a reg

    # --- Common fields ---
    short_title: str = ""
    long_title: str = ""
    formal_citation: str = ""        # composed differently for acts vs regulations
    pit_date: str = ""
    current_date: str = ""
    last_amended_date: str = ""
    inforce_start_date: str = ""
    source_url: str = ""
    parsed_at: str = ""

    # --- Statute-only ---
    consolidated_number: str = ""
    consolidated_official: bool = False
    annual_chapter: str = ""
    annual_year: str = ""
    annual_revised_statute: bool = False
    assented_date: str = ""
    consolidation_date: str = ""     # statute consolidation stage date (ISO)

    # --- Regulation-only ---
    instrument_number: str = ""       # "SOR/2021-181"
    regulation_type: str = ""         # "SOR", "SI", "CRC"
    gazette_part: str = ""            # "I" or "II"
    registration_date: str = ""       # ISO
    regulation_consolidation_date: str = ""  # ISO, from <ConsolidationDate>
    regulation_maker: str = ""        # "P.C." (Privy Council) etc.
    order_number: str = ""            # "2021-805"
    order_date: str = ""              # ISO
    enabling_authorities: list[EnablingAuthorityRow] = field(default_factory=list)

    # --- Body content (same shape for both types) ---
    headings: list[HeadingRow] = field(default_factory=list)
    provisions: list[ProvisionRow] = field(default_factory=list)
    sections: list[ProvisionRow] = field(default_factory=list)
    definitions: list[DefinitionRow] = field(default_factory=list)
    amendments: list[AmendmentRow] = field(default_factory=list)
    schedules: list["Schedule"] = field(default_factory=list)


@dataclass
class ScheduleItem:
    """One row of content from a schedule's list — a checklist-style data point.

    Justice Canada schedules use nested <List>/<Item> trees rather than the
    Section/Subsection hierarchy of the main body. We flatten them into rows,
    preserving nesting through an indent_level so the sheet keeps a visible
    structure (Items inside an outer Item like 15(a)/(b)/(c) get indent=1,
    and a third level like 15(a)(i) gets indent=2).
    """
    part_label: str = ""        # e.g. "PART A"
    part_title: str = ""        # e.g. "Information with Respect to ..."
    label: str = ""             # e.g. "1*" or "(a)" or "(i)"
    text: str = ""              # the item's body text
    indent_level: int = 0       # 0 = top-level item, 1 = sub-item, 2 = sub-sub-item
    in_force_date: str = ""
    last_amended_date: str = ""
    enacted_date: str = ""
    fid: str = ""
    repealed: bool = False
    starred: bool = False       # PCMLTFA-style "*" suffix marks mandatory fields


@dataclass
class Schedule:
    """A schedule appended to the end of an Act or Regulation (forms, tables,
    listed entities, etc.). One sheet per schedule in the workbook."""
    label: str = ""              # e.g. "SCHEDULE 1"
    title: str = ""              # e.g. "Suspicious Transaction or Attempted Transaction Report"
    originating_ref: str = ""    # e.g. "(Subsection 9(1) and section 11)"
    in_force_date: str = ""
    last_amended_date: str = ""
    fid: str = ""
    items: list[ScheduleItem] = field(default_factory=list)


# Backwards-compatible alias — internal references to ParsedAct still work.
ParsedAct = ParsedDocument


# ----------------------------------------------------------------------------
# XML loading
# ----------------------------------------------------------------------------

def load_xml(source: str) -> bytes:
    """Load XML bytes from a URL or local file path."""
    if source.startswith("http://") or source.startswith("https://"):
        req = urllib.request.Request(source, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.read()
    return Path(source).read_bytes()


# ----------------------------------------------------------------------------
# Library API — call from non-CLI consumers (Phase 2 ingest, tests, etc.)
# ----------------------------------------------------------------------------

def parse_document(source: str | Path | bytes,
                   *,
                   source_url: str | None = None) -> ParsedDocument:
    """Parse a Justice Laws XML document into a ``ParsedDocument``.

    ``source`` may be a URL, a file path, or the raw XML bytes. When raw
    bytes are passed, ``source_url`` is what gets recorded on the
    ``ParsedDocument`` so downstream renderers can build links back to
    the canonical text. URL/file callers normally don't need it — the
    function fills it in from the path.
    """
    # Resolve the source -> raw XML bytes.
    if isinstance(source, bytes):
        xml_bytes = source
        recorded_source = source_url or ""
    elif isinstance(source, Path):
        xml_bytes = source.read_bytes()
        recorded_source = source_url or f"file://{source.resolve()}"
    else:
        text = str(source)
        xml_bytes = load_xml(text)
        if text.startswith(("http://", "https://")):
            recorded_source = source_url or text
        else:
            recorded_source = source_url or f"file://{Path(text).resolve()}"

    parser = etree.XMLParser(remove_blank_text=False, recover=True)
    root = etree.fromstring(xml_bytes, parser=parser)

    parsed = ParsedDocument()
    parsed.source_url = recorded_source
    parsed.parsed_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Detect document type from the root element: <Statute> or
    # <Regulation>. Anything else falls back to statute parsing as a
    # best-effort default (matches main()'s behaviour).
    root_tag = etree.QName(root).localname
    parsed.document_type = "regulation" if root_tag == "Regulation" else "statute"

    parse_identification(root, parsed)
    parse_body(root, parsed)
    parse_schedules(root, parsed)
    return parsed


# ----------------------------------------------------------------------------
# Parsing
# ----------------------------------------------------------------------------

def parse_date_yyyy_mm_dd(date_elem) -> str:
    """Pull a YYYY-MM-DD string from a <Date><YYYY/><MM/><DD/></Date> element."""
    if date_elem is None:
        return ""
    y = inner_text(date_elem.find("YYYY")) or "0000"
    m = inner_text(date_elem.find("MM")) or "1"
    d = inner_text(date_elem.find("DD")) or "1"
    try:
        return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    except ValueError:
        return ""


def parse_identification(root, parsed: ParsedDocument) -> None:
    """Dispatch identification parsing based on the document root element.

    Statutes use <Chapter>/<ConsolidatedNumber>/<AnnualStatuteId>/<BillHistory>;
    regulations use <InstrumentNumber>/<RegistrationDate>/<ConsolidationDate>/
    <EnablingAuthority>/<RegulationMakerOrder>. The lims:* date attributes on
    the root element are common to both.
    """
    # The lims:* dates on the root element are common to both document types.
    parsed.pit_date = lims_attr(root, "pit-date")
    parsed.current_date = lims_attr(root, "current-date")
    parsed.last_amended_date = lims_attr(root, "lastAmendedDate")
    parsed.inforce_start_date = lims_attr(root, "inforce-start-date")

    if parsed.document_type == "regulation":
        parse_regulation_identification(root, parsed)
    else:
        parse_statute_identification(root, parsed)


def parse_statute_identification(root, parsed: ParsedDocument) -> None:
    ident = root.find("Identification")
    if ident is None:
        return
    parsed.long_title = inner_text(ident.find("LongTitle"))
    parsed.short_title = inner_text(ident.find("ShortTitle"))

    chapter = ident.find("Chapter")
    if chapter is not None:
        cons_elem = chapter.find("ConsolidatedNumber")
        parsed.consolidated_number = inner_text(cons_elem)
        # Per Justice Canada's LIMS2HTML.xsl: @official='yes' indicates the
        # citation that has legal evidentiary status (formal R.S.C. format);
        # @official='no' (or missing) means it's shown in parentheses as
        # informational only.
        parsed.consolidated_official = (
            cons_elem is not None and cons_elem.get("official", "no").lower() == "yes"
        )

        annual = chapter.find("AnnualStatuteId")
        if annual is not None:
            parsed.annual_chapter = inner_text(annual.find("AnnualStatuteNumber"))
            parsed.annual_year = inner_text(annual.find("YYYY"))
            parsed.annual_revised_statute = (
                annual.get("revised-statute", "no").lower() == "yes"
            )

    # Build a formal-style citation following the Justice Canada XSLT rules.
    parts: list[str] = []
    if parsed.annual_year and parsed.annual_chapter:
        prefix = "R.S.C." if parsed.annual_revised_statute else "S.C."
        parts.append(f"{prefix} {parsed.annual_year}, c. {parsed.annual_chapter}")
    if parsed.consolidated_official and parsed.consolidated_number:
        parts.append(f"R.S.C., 1985, c. {parsed.consolidated_number}")
    elif parsed.consolidated_number and not parts:
        # Unofficial — render as the bare chapter, consistent with the XSLT.
        parts.append(parsed.consolidated_number)
    parsed.formal_citation = "; ".join(parts)

    for stage in ident.findall(".//Stages"):
        date_str = parse_date_yyyy_mm_dd(stage.find("Date"))
        if not date_str:
            continue
        stage_name = stage.get("stage", "")
        if stage_name == "assented-to":
            parsed.assented_date = date_str
        elif stage_name == "consolidation":
            parsed.consolidation_date = date_str

    # Statutes use the consolidated chapter as the URL code.
    parsed.url_code = parsed.consolidated_number


def parse_regulation_identification(root, parsed: ParsedDocument) -> None:
    """Extract identification info from a <Regulation> root.

    Layout (per a SOR file from laws-lois.justice.gc.ca):
        <Regulation regulation-type="SOR" gazette-part="II" ...>
          <Identification>
            <InstrumentNumber>SOR/2021-181</InstrumentNumber>
            <RegistrationDate><Date>...</Date></RegistrationDate>
            <ConsolidationDate><Date>...</Date></ConsolidationDate>
            <EnablingAuthority>
              <XRefExternal reference-type="act" link="B-1.01">BANK ACT</XRefExternal>
              ...
            </EnablingAuthority>
            <LongTitle>...</LongTitle>
            <RegulationMakerOrder>
              <RegulationMaker>P.C.</RegulationMaker>
              <OrderNumber>2021-805</OrderNumber>
              <Date>...</Date>
            </RegulationMakerOrder>
          </Identification>
    """
    parsed.regulation_type = root.get("regulation-type", "")
    parsed.gazette_part = root.get("gazette-part", "")

    ident = root.find("Identification")
    if ident is None:
        return

    parsed.long_title = inner_text(ident.find("LongTitle"))
    parsed.short_title = inner_text(ident.find("ShortTitle"))  # often absent
    parsed.instrument_number = inner_text(ident.find("InstrumentNumber"))
    parsed.registration_date = parse_date_yyyy_mm_dd(
        ident.find("RegistrationDate/Date")
    )
    parsed.regulation_consolidation_date = parse_date_yyyy_mm_dd(
        ident.find("ConsolidationDate/Date")
    )

    enabling = ident.find("EnablingAuthority")
    if enabling is not None:
        for xref in enabling.findall("XRefExternal"):
            parsed.enabling_authorities.append(EnablingAuthorityRow(
                chapter_code=xref.get("link", ""),
                title=inner_text(xref),
                reference_type=xref.get("reference-type", ""),
            ))

    rmo = ident.find("RegulationMakerOrder")
    if rmo is not None:
        parsed.regulation_maker = inner_text(rmo.find("RegulationMaker"))
        parsed.order_number = inner_text(rmo.find("OrderNumber"))
        parsed.order_date = parse_date_yyyy_mm_dd(rmo.find("Date"))

    # The formal citation for a regulation is just its instrument number.
    parsed.formal_citation = parsed.instrument_number

    # URL code: turn "SOR/2021-181" → "SOR-2021-181" (URL-safe form).
    if parsed.instrument_number:
        parsed.url_code = parsed.instrument_number.replace("/", "-")
    elif parsed.regulation_type:
        # Fallback if InstrumentNumber is missing.
        parsed.url_code = parsed.regulation_type


def build_section_url(parsed: ParsedDocument, label: str) -> str:
    """Build a URL pointing at a specific section on the Justice Laws website,
    choosing the right URL pattern based on the document type."""
    if not (parsed.url_code and label):
        return ""
    pattern = (REGULATION_SECTION_URL if parsed.document_type == "regulation"
               else ACT_SECTION_URL)
    return pattern.format(code=parsed.url_code, label=label.strip())


def walk_schedule_list(list_elem, schedule: Schedule, part_label: str,
                        part_title: str, indent_level: int,
                        parent_label: str = "") -> None:
    """Recursively walk a <List> inside a Schedule, appending each <Item> as
    a flat row. Nested lists inside an item produce rows with higher indent
    AND a cumulative label that includes every ancestor's label — so an item
    (a) under item 14 lands in the sheet as ``14(a)`` rather than just ``(a)``,
    and a further-nested (i) under that becomes ``14(a)(i)``.

    The document-order traversal is important: an item with a nested list is
    emitted first, then its sub-items immediately follow.
    """
    for item in list_elem.findall("Item"):
        label_raw = inner_text(item.find("Label"))
        # PCMLTFA uses "1*", "6*" etc. — the "*" marks fields that must be
        # filled (rather than "where applicable"). Strip it from the label and
        # surface the fact in its own column so the sheet can be filtered on it.
        starred = label_raw.endswith("*")
        own_label = label_raw.rstrip("*") if starred else label_raw

        # Build the cumulative label by prepending every ancestor's label. For
        # top-level items there's no parent so this is just the item's own
        # label. For nested items the result threads through all ancestors:
        # 14 -> 14(a) -> 14(a)(i).
        cumulative_label = f"{parent_label}{own_label}" if parent_label else own_label

        text_elem = item.find("Text")
        text = inner_text(text_elem) if text_elem is not None else ""
        is_repealed = (text_elem is not None
                       and text_elem.find("Repealed") is not None)

        schedule.items.append(ScheduleItem(
            part_label=part_label,
            part_title=part_title,
            label=cumulative_label,
            text=text,
            indent_level=indent_level,
            in_force_date=lims_attr(item, "inforce-start-date"),
            last_amended_date=lims_attr(item, "lastAmendedDate"),
            enacted_date=lims_attr(item, "enacted-date"),
            fid=lims_attr(item, "fid"),
            repealed=is_repealed,
            starred=starred,
        ))

        # Nested lists inside this item become deeper-indented rows whose
        # labels carry this item's cumulative label as their prefix.
        for nested in item.findall("List"):
            walk_schedule_list(nested, schedule, part_label, part_title,
                               indent_level + 1, parent_label=cumulative_label)


def parse_schedules(root, parsed: ParsedDocument) -> None:
    """Find every <Schedule> element directly under the document root and
    convert it into a Schedule dataclass on ``parsed.schedules``.

    Schedules live as siblings of <Body> in the LIMS XML — they're not inside
    the Body. Each one carries its own ScheduleFormHeading (label, title,
    originating reference) and one or more <Heading>/<List> pairs (the Parts).
    """
    for schedule_elem in root.findall("Schedule"):
        schedule = Schedule()
        schedule.in_force_date = lims_attr(schedule_elem, "inforce-start-date")
        schedule.last_amended_date = lims_attr(schedule_elem, "lastAmendedDate")
        schedule.fid = lims_attr(schedule_elem, "fid")

        heading = schedule_elem.find("ScheduleFormHeading")
        if heading is not None:
            schedule.label = inner_text(heading.find("Label"))
            schedule.title = inner_text(heading.find("TitleText"))
            schedule.originating_ref = inner_text(heading.find("OriginatingRef"))

        # Walk Heading/List pairs in document order. A Heading sets the
        # "current Part" context; subsequent Lists (until the next Heading)
        # are attributed to that Part. Lists that appear before any Heading
        # carry an empty Part label, which is fine for schedules that don't
        # subdivide into parts.
        current_part_label = ""
        current_part_title = ""
        for child in schedule_elem:
            tag = etree.QName(child).localname
            if tag == "Heading":
                current_part_label = inner_text(child.find("Label"))
                current_part_title = inner_text(child.find("TitleText"))
            elif tag == "List":
                walk_schedule_list(child, schedule,
                                    current_part_label, current_part_title,
                                    indent_level=0)

        parsed.schedules.append(schedule)


def parse_body(root, parsed: ParsedAct) -> None:
    """Walk the <Body> in document order, tracking the current heading context."""
    body = root.find("Body")
    if body is None:
        return

    code = parsed.url_code
    current = {"PART": "", "DIVISION": "", "SUBDIVISION": ""}
    # Stack of plain (non-Part) heading titles by level — used for context only.
    heading_stack: dict[int, str] = {}

    for elem in body:
        tag = etree.QName(elem).localname
        if tag == "Heading":
            level = int(elem.get("level", "1"))
            label_text = inner_text(elem.find("Label"))
            title_text = inner_text(elem.find("TitleText"))

            parsed.headings.append(HeadingRow(
                level=level,
                label=label_text,
                title=title_text,
                in_force_date=lims_attr(elem, "inforce-start-date"),
                enacted_date=lims_attr(elem, "enacted-date"),
            ))

            # Track Parts by their explicit "PART 1" label; divisions/subdivisions
            # by level + title for the running context attached to each section.
            if label_text.upper().startswith("PART"):
                current["PART"] = f"{label_text}: {title_text}" if title_text else label_text
                current["DIVISION"] = ""
                current["SUBDIVISION"] = ""
                heading_stack.clear()
            else:
                heading_stack[level] = title_text
                # Clear deeper levels.
                for deeper in list(heading_stack):
                    if deeper > level:
                        heading_stack.pop(deeper, None)
                # Map levels: most acts use level 2 = Division-ish, level 3 = subdivision-ish.
                current["DIVISION"] = heading_stack.get(2, "")
                current["SUBDIVISION"] = heading_stack.get(3, "")

        elif tag == "Section":
            parse_section(elem, parsed, code, current.copy())


def parse_section(section_elem, parsed: ParsedAct, code: str, context: dict) -> None:
    section_label = inner_text(section_elem.find("Label"))
    marginal_note = inner_text(section_elem.find("MarginalNote"))
    in_force = lims_attr(section_elem, "inforce-start-date")
    last_amended = lims_attr(section_elem, "lastAmendedDate")
    enacted = lims_attr(section_elem, "enacted-date")
    fid = lims_attr(section_elem, "fid")
    section_url = build_section_url(parsed, section_label)

    section_text = collect_full_text(section_elem)
    # Detect a section that is just a "Repealed" marker (e.g. s. 13, s. 11.45).
    repealed_marker = section_elem.find("Text/Repealed")
    is_repealed = repealed_marker is not None and not any(
        section_elem.find(child) is not None
        for child in ("Subsection", "Paragraph")
    )

    section_citation = f"s. {section_label}" if section_label else ""

    # Append a row for the Section itself.
    section_row = ProvisionRow(
        citation=section_citation,
        section_url=section_url,
        level="Section",
        part=context["PART"],
        division=context["DIVISION"],
        subdivision=context["SUBDIVISION"],
        section=section_label,
        marginal_note=marginal_note,
        text=section_text,
        in_force_date=in_force,
        last_amended_date=last_amended,
        enacted_date=enacted,
        fid=fid,
        repealed=is_repealed,
    )
    parsed.provisions.append(section_row)
    parsed.sections.append(section_row)

    # Definitions can appear directly under a Section (rare) or under a Subsection.
    for definition_elem in section_elem.findall("Definition"):
        record_definition(definition_elem, parsed, section_citation, section_url)

    # Paragraphs may sit directly under the Section (no subsection wrapper).
    for paragraph_elem in section_elem.findall("Paragraph"):
        parse_paragraph(
            paragraph_elem, parsed, code, context, section_label,
            subsection_label="", section_url=section_url,
        )

    # Continued-section trailing text (after enumerated paragraphs) -- treat as
    # an addendum row anchored to the section.
    for cont in section_elem.findall("ContinuedSectionSubsection"):
        parsed.provisions.append(ProvisionRow(
            citation=section_citation,
            section_url=section_url,
            level="Continued",
            part=context["PART"],
            division=context["DIVISION"],
            subdivision=context["SUBDIVISION"],
            section=section_label,
            marginal_note=marginal_note,
            text=own_text(cont),
            in_force_date=lims_attr(cont, "inforce-start-date"),
            fid=lims_attr(cont, "fid"),
        ))

    # Subsections (the usual case for any section with multiple subsections).
    for subsection_elem in section_elem.findall("Subsection"):
        parse_subsection(subsection_elem, parsed, code, context, section_label, section_url)

    # HistoricalNote items become amendment rows.
    record_amendments(section_elem, parsed, section_citation, section_label)


def parse_subsection(sub_elem, parsed: ParsedAct, code: str, context: dict,
                     section_label: str, section_url: str) -> None:
    sub_label_raw = inner_text(sub_elem.find("Label"))   # e.g. "(1)" or "(2.1)"
    sub_label = sub_label_raw.strip()
    marginal_note = inner_text(sub_elem.find("MarginalNote"))
    citation = f"s. {section_label}{sub_label}"

    parsed.provisions.append(ProvisionRow(
        citation=citation,
        section_url=section_url,
        level="Subsection",
        part=context["PART"],
        division=context["DIVISION"],
        subdivision=context["SUBDIVISION"],
        section=section_label,
        subsection=sub_label,
        marginal_note=marginal_note,
        text=collect_full_text(sub_elem),
        in_force_date=lims_attr(sub_elem, "inforce-start-date"),
        last_amended_date=lims_attr(sub_elem, "lastAmendedDate"),
        enacted_date=lims_attr(sub_elem, "enacted-date"),
        fid=lims_attr(sub_elem, "fid"),
    ))

    for definition_elem in sub_elem.findall("Definition"):
        record_definition(definition_elem, parsed, citation, section_url)

    for paragraph_elem in sub_elem.findall("Paragraph"):
        parse_paragraph(
            paragraph_elem, parsed, code, context, section_label,
            subsection_label=sub_label, section_url=section_url,
        )

    for cont in sub_elem.findall("ContinuedSectionSubsection"):
        parsed.provisions.append(ProvisionRow(
            citation=citation,
            section_url=section_url,
            level="Continued",
            part=context["PART"],
            division=context["DIVISION"],
            subdivision=context["SUBDIVISION"],
            section=section_label,
            subsection=sub_label,
            text=own_text(cont),
            in_force_date=lims_attr(cont, "inforce-start-date"),
            fid=lims_attr(cont, "fid"),
        ))


def parse_paragraph(para_elem, parsed: ParsedAct, code: str, context: dict,
                    section_label: str, subsection_label: str, section_url: str) -> None:
    para_label = inner_text(para_elem.find("Label")).strip()
    citation = f"s. {section_label}{subsection_label}{para_label}"

    parsed.provisions.append(ProvisionRow(
        citation=citation,
        section_url=section_url,
        level="Paragraph",
        part=context["PART"],
        division=context["DIVISION"],
        subdivision=context["SUBDIVISION"],
        section=section_label,
        subsection=subsection_label,
        paragraph=para_label,
        text=collect_full_text(para_elem),
        in_force_date=lims_attr(para_elem, "inforce-start-date"),
        last_amended_date=lims_attr(para_elem, "lastAmendedDate"),
        enacted_date=lims_attr(para_elem, "enacted-date"),
        fid=lims_attr(para_elem, "fid"),
    ))

    for sub_para in para_elem.findall("Subparagraph"):
        parse_subparagraph(
            sub_para, parsed, code, context, section_label,
            subsection_label, para_label, section_url,
        )


def parse_subparagraph(subpara_elem, parsed: ParsedAct, code: str, context: dict,
                       section_label: str, subsection_label: str, paragraph_label: str,
                       section_url: str) -> None:
    sp_label = inner_text(subpara_elem.find("Label")).strip()
    citation = f"s. {section_label}{subsection_label}{paragraph_label}{sp_label}"

    parsed.provisions.append(ProvisionRow(
        citation=citation,
        section_url=section_url,
        level="Subparagraph",
        part=context["PART"],
        division=context["DIVISION"],
        subdivision=context["SUBDIVISION"],
        section=section_label,
        subsection=subsection_label,
        paragraph=paragraph_label,
        subparagraph=sp_label,
        text=collect_full_text(subpara_elem),
        in_force_date=lims_attr(subpara_elem, "inforce-start-date"),
        last_amended_date=lims_attr(subpara_elem, "lastAmendedDate"),
        enacted_date=lims_attr(subpara_elem, "enacted-date"),
        fid=lims_attr(subpara_elem, "fid"),
    ))

    for clause_elem in subpara_elem.findall("Clause"):
        cl_label = inner_text(clause_elem.find("Label")).strip()
        cl_citation = f"s. {section_label}{subsection_label}{paragraph_label}{sp_label}{cl_label}"
        parsed.provisions.append(ProvisionRow(
            citation=cl_citation,
            section_url=section_url,
            level="Clause",
            part=context["PART"],
            division=context["DIVISION"],
            subdivision=context["SUBDIVISION"],
            section=section_label,
            subsection=subsection_label,
            paragraph=paragraph_label,
            subparagraph=sp_label,
            clause=cl_label,
            text=collect_full_text(clause_elem),
            in_force_date=lims_attr(clause_elem, "inforce-start-date"),
            last_amended_date=lims_attr(clause_elem, "lastAmendedDate"),
            enacted_date=lims_attr(clause_elem, "enacted-date"),
            fid=lims_attr(clause_elem, "fid"),
        ))


def record_definition(definition_elem, parsed: ParsedAct,
                      citation: str, section_url: str) -> None:
    """Pull a Definition element apart and append a row to the Definitions sheet."""
    text_elem = definition_elem.find("Text")
    if text_elem is None:
        return
    term_en_elems = text_elem.findall(".//DefinedTermEn")
    term_fr_elems = text_elem.findall(".//DefinedTermFr")
    term_en = inner_text(term_en_elems[0]) if term_en_elems else ""
    term_fr = inner_text(term_fr_elems[0]) if term_fr_elems else ""
    repealed = text_elem.find("Repealed") is not None
    full_text = inner_text(definition_elem)

    parsed.definitions.append(DefinitionRow(
        term_en=term_en,
        term_fr=term_fr,
        citation=citation,
        section_url=section_url,
        text=full_text,
        in_force_date=lims_attr(definition_elem, "inforce-start-date"),
        enacted_date=lims_attr(definition_elem, "enacted-date"),
        repealed=repealed,
    ))


# Match the leading year inside a HistoricalNoteSubItem, e.g. "2024, c. 15, s. 280".
HIST_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def split_historical_note(text: str) -> list[str]:
    """Split a HistoricalNoteSubItem text on ';' the way Justice Canada's
    LIMS2HTML.xsl does (the recursive ``splitHN`` template). Each chunk is a
    discrete amendment citation, e.g.
        "2000, c. 17, s. 81; 2001, c. 41, s. 67; 2024, c. 15, s. 280"
    becomes three entries. Empty pieces are dropped.
    """
    if not text:
        return []
    return [chunk.strip() for chunk in text.split(";") if chunk.strip()]


def record_amendments(section_elem, parsed: ParsedAct,
                      section_citation: str, section_label: str) -> None:
    for hist in section_elem.findall("HistoricalNote"):
        for sub in hist.findall("HistoricalNoteSubItem"):
            full_text = inner_text(sub)
            parsed.amendments.append(AmendmentRow(
                citation=section_citation,
                section_label=section_label,
                in_force_date=lims_attr(sub, "inforce-start-date"),
                enacted_date=lims_attr(sub, "enacted-date"),
                enact_id=lims_attr(sub, "enactId"),
                text=full_text,
                citation_chunks=split_historical_note(full_text),
            ))


# ----------------------------------------------------------------------------
# Optional: resolve XRefExternal link codes (e.g. "C-46" -> "Criminal Code")
# ----------------------------------------------------------------------------
# Justice Canada publishes a lookup directory at
#     https://github.com/justicecanada/laws-lois-xml/tree/main/lookup
# that maps chapter codes used in <XRefExternal link="..."> to act and
# regulation titles. The current build pipeline doesn't bundle a copy because
# the directory isn't reachable from this script's typical run environment,
# but if you clone that repo locally you can populate LOOKUP_TABLE below by
# parsing the files under /lookup, then call resolve_external_link("C-46")
# to get the title back. This is left as an opt-in hook so the script still
# works offline / without the repo.
LOOKUP_TABLE: dict[str, str] = {}


def resolve_external_link(code: str) -> str:
    """Return the full title for a chapter code, or the code itself if unknown."""
    if not code:
        return ""
    return LOOKUP_TABLE.get(code.strip(), code.strip())


def load_lookup_from_directory(path: Path) -> None:
    """Populate LOOKUP_TABLE from a local clone of justicecanada/laws-lois-xml/lookup.

    Each file there is expected to contain a mapping between chapter codes
    and titles. Adapt the parser below once you've inspected the actual file
    format on disk; the stub is written defensively so it won't crash on
    unexpected content.
    """
    if not path.exists():
        return
    for entry in path.iterdir():
        if not entry.is_file():
            continue
        try:
            content = entry.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        # Defensive parsing: handles "CODE\tTITLE" and "CODE,TITLE" line formats.
        for line in content.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            for sep in ("\t", ","):
                if sep in line:
                    code, _, title = line.partition(sep)
                    if code and title:
                        LOOKUP_TABLE[code.strip()] = title.strip()
                    break


# ----------------------------------------------------------------------------
# Excel rendering
# ----------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="000000")  # black with white bold text
# Note: the yellow recent-amendment fill and pink repealed fill are applied
# via Excel conditional formatting rules rather than direct cell fills, so
# they live inside add_recent_and_repealed_conditional_formatting().
THIN = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 80) -> None:
    """Best-effort column width based on longest cell value (text columns get wrap)."""
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        longest = 0
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            s = str(value)
            longest = max(longest, min(len(s), max_width))
        ws.column_dimensions[letter].width = max(12, longest + 2)


def write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    ws.freeze_panes = "A2"


# Format codes used everywhere dates are stored as real Excel values.
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


def to_acronym(s: str) -> str:
    """Build an acronym from the first letter of each word in a title, used
    for output filenames. The word "Regulations" is excluded (case-insensitive)
    because it appears at the end of every regulation name and would
    otherwise pad the acronym with a redundant 'R' for every regulation.

    Examples:
        "Bank Act"                                         -> "BA"
        "Cooperative Credit Associations Act"              -> "CCAA"
        "Financial Consumer Protection Framework Regulations"
                                                           -> "FCPF"
        "Proceeds of Crime (Money Laundering) and Terrorist Financing Act"
                                                           -> "POCMLATFA"
    """
    if not s:
        return ""
    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", s)
    words = cleaned.split()
    return "".join(
        w[0].upper() for w in words
        if w and w.lower() != "regulations"
    )


def to_excel_date(s):
    """Convert an ISO date string ('YYYY-MM-DD') into a ``datetime.date`` so
    Excel treats it as a real date (sortable, filterable, usable in formulas).

    Returns:
        - ``datetime.date`` for a parseable ISO date,
        - ``""`` for empty input (so the cell is genuinely blank),
        - the original string for anything that doesn't parse, so unexpected
          values stay visible rather than being silently dropped.
    """
    if not s:
        return ""
    try:
        return dt.date.fromisoformat(s)
    except (ValueError, TypeError):
        return s


def to_excel_datetime(s):
    """Convert a 'YYYY-MM-DD HH:MM:SS' (or plain ISO date) string into a
    ``datetime.datetime``. Same fallback behaviour as ``to_excel_date``.
    """
    if not s:
        return ""
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        pass
    try:
        return dt.datetime.fromisoformat(s)
    except (ValueError, TypeError):
        return s


def set_column_date_format(ws, col_letters, fmt: str = DATE_FORMAT) -> None:
    """Apply a date number-format to every data cell in the given columns."""
    for col_letter in col_letters:
        for cell in ws[col_letter]:
            if cell.row > 1:  # skip header row
                cell.number_format = fmt


def write_readme(wb, parsed: ParsedDocument) -> None:
    ws = wb.active
    ws.title = "README"

    tracking_paragraph = (
        "Re-run the script periodically and compare the Amendments sheet "
        "against your last run. The lims:enactId is a stable identifier for "
        "each amendment instrument, so you can diff cleanly. Provisions with "
        "a Last Amended date in the past year are highlighted in yellow; "
        "repealed provisions in pink. Both are applied via Excel conditional "
        "formatting rules, so they recompute automatically as time passes — "
        "no need to re-run the script just to refresh the highlights."
    )

    is_reg = parsed.document_type == "regulation"
    sections: list[tuple[str, object]] = [
        ("Justice Laws extract", ""),
        ("Document type", parsed.document_type.capitalize()),
        ("Generated", to_excel_datetime(parsed.parsed_at)),
        ("Source", parsed.source_url),
        ("Short title", parsed.short_title or parsed.long_title),
    ]
    if is_reg:
        sections.extend([
            ("Instrument number", parsed.instrument_number),
            ("Regulation type", parsed.regulation_type),
            ("Consolidation date", to_excel_date(parsed.regulation_consolidation_date)),
        ])
    else:
        sections.extend([
            ("Chapter", parsed.consolidated_number),
            ("Annual citation",
             f"{parsed.annual_year}, c. {parsed.annual_chapter}" if parsed.annual_chapter else ""),
            ("Consolidation date", to_excel_date(parsed.consolidation_date)),
        ])
    sections.extend([
        ("Last amended (per source)", to_excel_date(parsed.last_amended_date)),
        ("Point-in-time as of", to_excel_date(parsed.pit_date)),
        ("", ""),
        ("How to use this workbook", ""),
        ("  Metadata", "Document-level facts and the URL the workbook was built from."),
    ])
    if is_reg:
        sections.append(("  EnablingAuthorities",
                         "The Acts of Parliament that authorize this regulation."))
    sections.extend([
        ("  Structure", "Outline of Parts, Divisions, and other headings."),
        ("  Provisions", "Every level of every section as one row, with built citations."),
        ("  Sections", "One row per Section — fastest sheet for marginal-note lookup."),
        ("  Definitions", "Defined terms (EN/FR) with the section they're defined in."),
        ("  Amendments", "HistoricalNote entries split into Most-recent vs Prior citations."),
        ("  Schedule N", "One sheet per schedule appended to the act (e.g. report forms)."),
        ("  Search", "Live keyword search across all provisions (type in cell B5)."),
        ("", ""),
        ("Tracking changes over time", tracking_paragraph),
    ])

    for row_idx, (k, v) in enumerate(sections, start=1):
        ws.cell(row=row_idx, column=1, value=k).font = Font(bold=k and not k.startswith("  "))
        value_cell = ws.cell(row=row_idx, column=2, value=v)
        if isinstance(v, dt.datetime):
            value_cell.number_format = DATETIME_FORMAT
        elif isinstance(v, dt.date):
            value_cell.number_format = DATE_FORMAT
        # The "Tracking changes" paragraph is long — let it wrap, and give
        # the row enough height that the wrapped lines are visible without
        # the user having to drag the row border.
        if k == "Tracking changes over time":
            value_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True,
            )
            ws.row_dimensions[row_idx].height = 105
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100


def write_metadata(wb, parsed: ParsedDocument) -> None:
    ws = wb.create_sheet("Metadata")
    rows: list[tuple[str, object]] = [
        ("Field", "Value"),
        ("Document type", parsed.document_type.capitalize()),
        ("Short title", parsed.short_title),
        ("Long title", parsed.long_title),
        ("Formal citation", parsed.formal_citation),
    ]
    if parsed.document_type == "regulation":
        rows.extend([
            ("Instrument number", parsed.instrument_number),
            ("Regulation type", parsed.regulation_type),
            ("Gazette part", parsed.gazette_part),
            ("Registration date", to_excel_date(parsed.registration_date)),
            ("Consolidation date", to_excel_date(parsed.regulation_consolidation_date)),
            ("Regulation maker", parsed.regulation_maker),
            ("Order number", parsed.order_number),
            ("Order date", to_excel_date(parsed.order_date)),
        ])
    else:
        rows.extend([
            ("Consolidated chapter", parsed.consolidated_number),
            ("Consolidated chapter is official", "Yes" if parsed.consolidated_official else "No"),
            ("Annual statute chapter", parsed.annual_chapter),
            ("Annual statute year", parsed.annual_year),
            ("Annual statute is revised (R.S.C.)", "Yes" if parsed.annual_revised_statute else "No"),
            ("Date assented to", to_excel_date(parsed.assented_date)),
            ("Consolidation date", to_excel_date(parsed.consolidation_date)),
        ])
    rows.extend([
        ("Point-in-time date", to_excel_date(parsed.pit_date)),
        ("Current as of", to_excel_date(parsed.current_date)),
        ("Last amended date", to_excel_date(parsed.last_amended_date)),
        ("In-force start date", to_excel_date(parsed.inforce_start_date)),
        ("Source URL", parsed.source_url),
        ("Workbook generated", to_excel_datetime(parsed.parsed_at)),
    ])
    for row in rows:
        ws.append(row)
    style_header(ws, 2)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 110
    # Format the value cell appropriately based on its Python type. dt.datetime
    # must be checked before dt.date because datetime is a subclass of date.
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        if isinstance(cell.value, dt.datetime):
            cell.number_format = DATETIME_FORMAT
        elif isinstance(cell.value, dt.date):
            cell.number_format = DATE_FORMAT
    # Make the URL clickable.
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Source URL":
            cell = ws.cell(row=r, column=2)
            cell.hyperlink = parsed.source_url
            cell.font = Font(color="0563C1", underline="single")


def write_structure(wb, parsed: ParsedAct) -> None:
    ws = wb.create_sheet("Structure")
    headers = ["Heading level", "Label", "Title", "In-force date", "Enacted date"]
    write_header(ws, headers)
    for h in parsed.headings:
        ws.append([
            h.level, h.label, h.title,
            to_excel_date(h.in_force_date), to_excel_date(h.enacted_date),
        ])
    set_column_date_format(ws, ["D", "E"])  # In-force date, Enacted date
    add_alternating_row_banding(ws, n_data_rows=len(parsed.headings), n_cols=len(headers))
    autosize(ws)


def write_provisions(wb, parsed: ParsedDocument) -> None:
    ws = wb.create_sheet("Provisions")
    # Context is a combined column placed right after Level. It joins
    # Marginal note, Subdivision, Division, Part (in that order, skipping
    # empty values) with commas, so a single column gives the reader the
    # full structural location of the provision at a glance. The individual
    # Part/Division/Subdivision/Section/.../Marginal-note columns are kept
    # in the sheet but hidden by default — they're available for filtering
    # if needed, but Context is what people read.
    headers = [
        "Citation", "Level", "Context",
        "Part", "Division", "Subdivision",
        "Section", "Subsection", "Paragraph", "Subparagraph", "Clause",
        "Marginal note", "Text",
        "In-force date", "Last amended", "Enacted date",
        "Repealed", "lims:fid", "Source URL",
    ]
    write_header(ws, headers)

    # For sub-level provisions, the row's own marginal_note is usually empty
    # (subsections/paragraphs rarely carry their own marginal note). Track the
    # most recent Section-level marginal note as we iterate so the Context
    # string for sub-rows still picks up the parent section's heading.
    current_section_marginal = ""
    for p in parsed.provisions:
        if p.level == "Section":
            current_section_marginal = p.marginal_note
        marginal_for_context = p.marginal_note or current_section_marginal
        context_str = ", ".join(
            x for x in (marginal_for_context, p.subdivision, p.division, p.part) if x
        )
        ws.append([
            p.citation, p.level, context_str,
            p.part, p.division, p.subdivision,
            p.section, p.subsection, p.paragraph, p.subparagraph, p.clause,
            p.marginal_note, p.text,
            to_excel_date(p.in_force_date),
            to_excel_date(p.last_amended_date),
            to_excel_date(p.enacted_date),
            "Yes" if p.repealed else "", p.fid, p.section_url,
        ])

    # Date columns shifted by one due to the new Context column at position 3.
    set_column_date_format(ws, ["N", "O", "P"])  # In-force, Last amended, Enacted
    apply_url_hyperlinks(ws, url_col=19)
    add_recent_and_repealed_conditional_formatting(
        ws,
        n_data_rows=len(parsed.provisions),
        n_cols=len(headers),
        date_col_letter="O",      # "Last amended"  (was N before Context column)
        repealed_col_letter="Q",  # "Repealed"      (was P before Context column)
    )
    autosize(ws, max_width=70)

    # Hide the columns from Part through Marginal note — they're useful for
    # advanced filtering but Context already presents the high-level location
    # and Citation already presents the section/subsection/paragraph hierarchy.
    for col_letter in ("D", "E", "F", "G", "H", "I", "J", "K", "L"):
        ws.column_dimensions[col_letter].hidden = True

    # Give the Context and Text columns generous widths and turn on wrapping.
    ws.column_dimensions["C"].width = 60     # Context
    ws.column_dimensions["M"].width = 90     # Text
    for col_letter in ("C", "M"):
        for row in ws.iter_rows(
            min_row=2,
            min_col=ws[col_letter + "1"].column,
            max_col=ws[col_letter + "1"].column,
        ):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions


def write_sections(wb, parsed: ParsedAct) -> None:
    ws = wb.create_sheet("Sections")
    headers = [
        "Citation", "Section", "Marginal note", "Part", "Division",
        "In-force date", "Last amended", "Source URL",
    ]
    write_header(ws, headers)
    for p in parsed.sections:
        ws.append([
            p.citation, p.section, p.marginal_note, p.part, p.division,
            to_excel_date(p.in_force_date), to_excel_date(p.last_amended_date),
            p.section_url,
        ])
    set_column_date_format(ws, ["F", "G"])  # In-force date, Last amended
    apply_url_hyperlinks(ws, url_col=8)
    add_recent_and_repealed_conditional_formatting(
        ws,
        n_data_rows=len(parsed.sections),
        n_cols=len(headers),
        date_col_letter="G",      # "Last amended" column
        repealed_col_letter=None, # Sections sheet doesn't have a Repealed column
    )
    autosize(ws, max_width=60)
    ws.auto_filter.ref = ws.dimensions


def write_definitions(wb, parsed: ParsedAct) -> None:
    ws = wb.create_sheet("Definitions")
    headers = [
        "Term (EN)", "Term (FR)", "Citation", "Definition text",
        "In-force date", "Enacted date", "Repealed", "Source URL",
    ]
    write_header(ws, headers)
    for d in parsed.definitions:
        ws.append([
            d.term_en, d.term_fr, d.citation, d.text,
            to_excel_date(d.in_force_date), to_excel_date(d.enacted_date),
            "Yes" if d.repealed else "", d.section_url,
        ])
    set_column_date_format(ws, ["E", "F"])  # In-force date, Enacted date
    apply_url_hyperlinks(ws, url_col=8)
    add_recent_and_repealed_conditional_formatting(
        ws,
        n_data_rows=len(parsed.definitions),
        n_cols=len(headers),
        date_col_letter=None,     # No "last amended" tracked on definitions
        repealed_col_letter="G",  # "Repealed" column
    )
    autosize(ws, max_width=70)
    ws.column_dimensions["D"].width = 100
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions


def write_amendments(wb, parsed: ParsedAct) -> None:
    ws = wb.create_sheet("Amendments")
    headers = [
        "Section", "Citation", "In-force date", "Enacted date",
        "Enact ID", "Most recent citation", "Prior citations", "Amendment text",
    ]
    write_header(ws, headers)
    # Sort newest-first by in-force date so recent amendments are easy to find.
    sorted_amendments = sorted(
        parsed.amendments,
        key=lambda a: (a.in_force_date or "0000-00-00"),
        reverse=True,
    )
    for a in sorted_amendments:
        # Per the Justice Canada XSLT, semicolon-separated chunks within one
        # HistoricalNoteSubItem are listed in chronological order, so the LAST
        # chunk is the most recent citation (matching lims:inforce-start-date).
        most_recent = a.citation_chunks[-1] if a.citation_chunks else a.text
        prior = "; ".join(a.citation_chunks[:-1]) if len(a.citation_chunks) > 1 else ""
        ws.append([
            a.section_label, a.citation,
            to_excel_date(a.in_force_date), to_excel_date(a.enacted_date),
            a.enact_id, most_recent, prior, a.text,
        ])
    set_column_date_format(ws, ["C", "D"])  # In-force date, Enacted date
    # Highlight rows whose in-force date falls within the past year (live rule).
    add_recent_and_repealed_conditional_formatting(
        ws,
        n_data_rows=len(sorted_amendments),
        n_cols=len(headers),
        date_col_letter="C",      # "In-force date" column
        repealed_col_letter=None, # Amendments aren't themselves repealed
    )
    autosize(ws, max_width=80)
    # Widen the wrap-text columns: Most recent (F), Prior (G), Amendment text (H).
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["H"].width = 90
    for col_letter in ("F", "G", "H"):
        for row in ws.iter_rows(min_row=2, min_col=ws[col_letter + "1"].column,
                                max_col=ws[col_letter + "1"].column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions


def add_recent_and_repealed_conditional_formatting(
    ws,
    n_data_rows: int,
    n_cols: int,
    date_col_letter: str | None = None,
    repealed_col_letter: str | None = None,
) -> None:
    """Add two conditional-formatting rules to a worksheet:

    1. Pink highlight (#FFCCFF) when the Repealed column equals "Yes".
    2. Yellow highlight (#FFFF99) when the date column holds a date within
       the past year (recomputed live via TODAY() in Excel).

    The repealed rule is added first so it takes precedence over the recent
    rule when both would otherwise apply (matching the previous behaviour
    where the repealed colour overwrote the recent colour). Rules use
    anchored column references ($-prefixed), so the formatting follows the
    whole row across all columns.

    Date columns must hold real Excel date values (numbers) rather than
    ISO strings — see ``to_excel_date()`` and ``set_column_date_format()``.
    """
    if n_data_rows < 1:
        return
    last_row = n_data_rows + 1  # +1 for the header row
    full_range = f"A2:{get_column_letter(n_cols)}{last_row}"

    if repealed_col_letter:
        pink_fill = PatternFill("solid", start_color="FFCCFF", end_color="FFCCFF")
        ws.conditional_formatting.add(
            full_range,
            FormulaRule(
                formula=[f'${repealed_col_letter}2="Yes"'],
                fill=pink_fill,
                stopIfTrue=True,
            ),
        )

    if date_col_letter:
        yellow_fill = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
        # The date column holds real Excel date values (see to_excel_date()), so
        # we can do straightforward date arithmetic. ISNUMBER guards against
        # blank cells and any string fallbacks for unparseable input.
        ws.conditional_formatting.add(
            full_range,
            FormulaRule(
                formula=[
                    f'AND(ISNUMBER(${date_col_letter}2),'
                    f'${date_col_letter}2>=TODAY()-365)'
                ],
                fill=yellow_fill,
            ),
        )

    # Alternating row banding — added LAST so it has the lowest priority
    # (the pink and yellow rules above override it on matching rows).
    banding_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=banding_fill),
    )


def add_alternating_row_banding(ws, n_data_rows: int, n_cols: int) -> None:
    """Standalone banding rule for sheets that don't use the recent/repealed
    helper (Structure, EnablingAuthorities, Search). Adds a single
    conditional-formatting rule that fills even-numbered rows with a light grey.
    """
    if n_data_rows < 1:
        return
    last_row = n_data_rows + 1  # +1 for the header row
    full_range = f"A2:{get_column_letter(n_cols)}{last_row}"
    banding_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=banding_fill),
    )


def apply_url_hyperlinks(ws, url_col: int | None) -> None:
    """Turn a column of URLs into clickable hyperlinks (one-time, not a rule)."""
    if not url_col:
        return
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=url_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="0563C1", underline="single")


def write_enabling_authorities(wb, parsed: ParsedDocument) -> None:
    """Regulations are made under one or more enabling Acts of Parliament.
    The Justice Laws XML lists these in <EnablingAuthority>. We surface them
    on their own sheet because a single regulation can have many parents
    (the Financial Consumer Protection Framework Regulations, for instance,
    cite five enabling Acts) — too many for the Metadata key/value layout.
    """
    if not parsed.enabling_authorities:
        return
    ws = wb.create_sheet("EnablingAuthorities")
    headers = ["Chapter code", "Title", "Reference type", "Justice Laws URL"]
    write_header(ws, headers)
    for ea in parsed.enabling_authorities:
        # The XML's link attribute is the same chapter code used in the
        # Justice Laws website URL for that act.
        if ea.chapter_code and ea.reference_type == "act":
            url = f"https://laws-lois.justice.gc.ca/eng/acts/{ea.chapter_code}/"
        elif ea.chapter_code and ea.reference_type == "regulation":
            url = f"https://laws-lois.justice.gc.ca/eng/regulations/{ea.chapter_code}/"
        else:
            url = ""
        ws.append([ea.chapter_code, ea.title, ea.reference_type, url])
    apply_url_hyperlinks(ws, url_col=4)
    add_alternating_row_banding(ws,
                                n_data_rows=len(parsed.enabling_authorities),
                                n_cols=len(headers))
    autosize(ws, max_width=80)
    ws.column_dimensions["B"].width = 80
    ws.auto_filter.ref = ws.dimensions


def _safe_sheet_name(raw: str, fallback: str) -> str:
    """Clean an arbitrary string into a valid Excel sheet name.

    Rules: max 31 chars, may not contain : \\ / ? * [ ]. Returns the fallback
    if the cleaned name comes out empty.
    """
    if not raw:
        return fallback
    # "SCHEDULE 1" -> "Schedule 1" reads better in the tab bar.
    cleaned = raw.title() if raw.isupper() else raw
    for ch in ':\\/?*[]':
        cleaned = cleaned.replace(ch, "")
    cleaned = cleaned.strip()[:31]
    return cleaned or fallback


def write_schedules(wb, parsed: ParsedDocument) -> None:
    """Add one sheet per Schedule. Layout:

        Row 1     SCHEDULE N                                  (black banner)
        Row 2     <Title of the schedule>                     (bold)
        Row 3     (originating reference)                     (italic)
        Row 4     <blank>
        Row 5     Parts:                                      (bold)
        Row 6..k  PART A | Information with Respect to ...    (legend rows)
                  PART B | Information with Respect to ...
                  ...
        Row k+1   <blank>
        Row k+2   [headers]
        Row k+3+  data rows

    Pulling the part titles into a legend at the top keeps the data table
    narrower (no Part-title column) and gives the reader a single place to
    learn what each PART covers without having to scroll. Nested items are
    indented visually AND carry cumulative labels (``14(a)``, ``14(a)(i)``),
    so the hierarchy is obvious both at a glance and when the sheet is
    filtered/sorted out of document order.
    """
    if not parsed.schedules:
        return

    for idx, sch in enumerate(parsed.schedules, start=1):
        sheet_name = _safe_sheet_name(sch.label, f"Schedule {idx}")
        # Disambiguate if another sheet by the same name already exists.
        original = sheet_name
        counter = 2
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original[:28]} ({counter})"
            counter += 1
        ws = wb.create_sheet(sheet_name)

        # Nine data columns now (Part title removed in favour of the legend).
        headers = [
            "Part", "Label", "Mandatory", "Text",
            "In-force date", "Last amended", "Enacted date",
            "Repealed", "lims:fid",
        ]
        n_cols = len(headers)
        last_col_letter = get_column_letter(n_cols)

        # ---- Row 1: black banner with the schedule label ----
        ws["A1"] = sch.label or f"Schedule {idx}"
        ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", start_color="000000")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = 32

        # ---- Row 2: descriptive title under the banner ----
        if sch.title:
            ws["A2"] = sch.title
            ws["A2"].font = Font(name="Calibri", size=12, bold=True)
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
            ws.row_dimensions[2].height = 22

        # ---- Row 3: originating reference (italic) ----
        if sch.originating_ref:
            ws["A3"] = sch.originating_ref
            ws["A3"].font = Font(name="Calibri", size=10, italic=True)
            ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)

        # ---- Rows 5..k: parts legend ----
        # Collect unique parts in document order. dict preserves insertion
        # order in Python 3.7+ so this gives us a deduped, ordered list of
        # (label, title) pairs without needing collections.OrderedDict.
        parts_seen: dict[str, str] = {}
        for item in sch.items:
            if item.part_label and item.part_label not in parts_seen:
                parts_seen[item.part_label] = item.part_title

        if parts_seen:
            ws.cell(row=5, column=1, value="Parts:").font = Font(bold=True, size=11)
            legend_row = 6
            for plabel, ptitle in parts_seen.items():
                ws.cell(row=legend_row, column=1, value=plabel).font = (
                    Font(bold=True, size=10)
                )
                ws.cell(row=legend_row, column=2, value=ptitle).font = (
                    Font(size=10, italic=True)
                )
                ws.merge_cells(start_row=legend_row, start_column=2,
                               end_row=legend_row, end_column=n_cols)
                legend_row += 1
            headers_row = legend_row + 1     # one blank row after the legend
        else:
            headers_row = 5                  # no legend, headers go where they used to

        # ---- Data headers ----
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=headers_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[headers_row].height = 22

        # ---- Data rows ----
        data_start_row = headers_row + 1
        for item in sch.items:
            ws.append([
                item.part_label,
                item.label,                  # already cumulative (e.g. "14(a)")
                "Yes" if item.starred else "",
                item.text,
                to_excel_date(item.in_force_date),
                to_excel_date(item.last_amended_date),
                to_excel_date(item.enacted_date),
                "Yes" if item.repealed else "",
                item.fid,
            ])

        # ---- Visual indent on the Label column for nested items ----
        # The cumulative label already conveys the hierarchy textually, but
        # indenting reinforces it visually so the eye can scan structure at
        # a glance without reading each label.
        n_items = len(sch.items)
        if n_items:
            label_col = 2  # "Label" is now column B (was C with Part title present)
            for i, item in enumerate(sch.items):
                row = data_start_row + i
                if item.indent_level > 0:
                    ws.cell(row=row, column=label_col).alignment = Alignment(
                        horizontal="left", vertical="top",
                        indent=item.indent_level * 2,
                    )

        # Date columns: E, F, G (shifted left by one after Part-title removal).
        set_column_date_format(ws, ["E", "F", "G"])

        # ---- Conditional formatting on the data range ----
        # Pink for repealed (column H now), banding for everything else. The
        # formula references the first data row so Excel calculates the
        # relative offset for each row correctly.
        if n_items > 0:
            last_data_row = data_start_row + n_items - 1
            data_range = f"A{data_start_row}:{last_col_letter}{last_data_row}"
            pink_fill = PatternFill("solid", start_color="FFCCFF",
                                     end_color="FFCCFF")
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(formula=[f'$H{data_start_row}="Yes"'],
                            fill=pink_fill, stopIfTrue=True),
            )
            banding_fill = PatternFill("solid", start_color="F2F2F2",
                                        end_color="F2F2F2")
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(formula=["MOD(ROW(),2)=0"], fill=banding_fill),
            )

        # ---- Column widths ----
        ws.column_dimensions["A"].width = 10   # Part
        ws.column_dimensions["B"].width = 14   # Label (fits "14(a)(iii)")
        ws.column_dimensions["C"].width = 11   # Mandatory
        ws.column_dimensions["D"].width = 90   # Text
        ws.column_dimensions["E"].width = 14   # In-force
        ws.column_dimensions["F"].width = 14   # Last amended
        ws.column_dimensions["G"].width = 14   # Enacted
        ws.column_dimensions["H"].width = 11   # Repealed
        ws.column_dimensions["I"].width = 12   # lims:fid

        # Wrap the Text column.
        for r in range(data_start_row, data_start_row + n_items):
            ws.cell(row=r, column=4).alignment = Alignment(
                wrap_text=True, vertical="top",
            )

        # AutoFilter on the data table.
        if n_items > 0:
            ws.auto_filter.ref = f"A{headers_row}:{last_col_letter}{data_start_row + n_items - 1}"

        # Freeze everything above the first data row so the banner, legend,
        # and headers stay visible while scrolling.
        ws.freeze_panes = f"A{data_start_row}"


def write_search(wb, parsed: ParsedDocument) -> None:
    """Search sheet: static Provisions data with live conditional highlighting.

    Why no FILTER: openpyxl writes dynamic-array functions (FILTER, SORT …)
    using the old CSE array-formula XML attributes which are structurally
    wrong for Excel 365 spill formulas and cause "Removed Records: Formula"
    errors in many Excel versions. This implementation avoids all dynamic-array
    functions and instead:

      - Writes every Provisions row directly to the sheet (static data).
      - Uses a SUMPRODUCT formula at E5 for the match count (works everywhere).
      - Uses a conditional-formatting rule to highlight matching rows in yellow.
        The CF formula evaluates per-row against the input cell $B$5, so the
        highlight updates the moment the user types — no volatile helper column
        needed, no spill formula needed.
      - Enables AutoFilter so the user can immediately filter to highlighted
        rows via Home → AutoFilter → Filter by Color → yellow.
    """
    ws = wb.create_sheet("Search")
    n = len(parsed.provisions)
    if n == 0:
        ws["A1"] = "(No provisions to search.)"
        return

    # ---- Row 1: black banner ----
    ws["A1"] = "Search Provisions"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="000000")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A1:S1")
    ws.row_dimensions[1].height = 32

    # ---- Row 3: instructions ----
    ws["A3"] = (
        "Type a keyword, phrase, or section number in the yellow cell below. "
        "Matching rows highlight yellow as you type (searches Citation, Context, "
        "Marginal note and Text). Match count appears in E5. "
        "To show only matches: AutoFilter > Filter by Color > yellow."
    )
    ws["A3"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("A3:S3")
    ws.row_dimensions[3].height = 36

    # ---- Row 5: search input + match count ----
    ws["A5"] = "Search:"
    ws["A5"].font = Font(bold=True)
    ws["A5"].alignment = Alignment(horizontal="right", vertical="center")

    input_cell = ws["B5"]
    input_cell.fill = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
    input_cell.font = Font(size=12, bold=True)
    input_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    thin_dark = Side(border_style="medium", color="000000")
    input_cell.border = Border(
        left=thin_dark, right=thin_dark, top=thin_dark, bottom=thin_dark,
    )
    ws.merge_cells("B5:C5")
    ws.row_dimensions[5].height = 26

    ws["D5"] = "Matches:"
    ws["D5"].font = Font(bold=True)
    ws["D5"].alignment = Alignment(horizontal="right", vertical="center")

    # ---- Row 7: column headers (same order as Provisions) ----
    headers = [
        "Citation", "Level", "Context",
        "Part", "Division", "Subdivision",
        "Section", "Subsection", "Paragraph", "Subparagraph", "Clause",
        "Marginal note", "Text",
        "In-force date", "Last amended", "Enacted date",
        "Repealed", "lims:fid", "Source URL",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 24

    # ---- Rows 8+: static Provisions data ----
    # Written directly (no formula spill) so there is no dependency on
    # dynamic-array support. The same Context-inheritance logic as Provisions.
    current_section_marginal = ""
    for p in parsed.provisions:
        if p.level == "Section":
            current_section_marginal = p.marginal_note
        marginal_for_context = p.marginal_note or current_section_marginal
        context_str = ", ".join(
            x for x in (marginal_for_context, p.subdivision, p.division, p.part)
            if x
        )
        ws.append([
            p.citation, p.level, context_str,
            p.part, p.division, p.subdivision,
            p.section, p.subsection, p.paragraph, p.subparagraph, p.clause,
            p.marginal_note, p.text,
            to_excel_date(p.in_force_date),
            to_excel_date(p.last_amended_date),
            to_excel_date(p.enacted_date),
            "Yes" if p.repealed else "", p.fid, p.section_url,
        ])

    last_data_row = 7 + n      # row 7 = headers; rows 8 … 7+n = data

    # ---- E5: match count (SUMPRODUCT, no dynamic arrays) ----
    # References this sheet's own columns A, C, L, M (same layout as Provisions).
    # Using LEN($B$5)>0 guard avoids counting everything when the cell is empty.
    ws["E5"] = (
        f'=IF(LEN($B$5)=0,"",'
        f'SUMPRODUCT(--('
        f'(ISNUMBER(SEARCH($B$5,A8:A{last_data_row}))+'
        f'ISNUMBER(SEARCH($B$5,C8:C{last_data_row}))+'
        f'ISNUMBER(SEARCH($B$5,L8:L{last_data_row}))+'
        f'ISNUMBER(SEARCH($B$5,M8:M{last_data_row})))>0)))'
    )
    ws["E5"].font = Font(bold=True)
    ws["E5"].alignment = Alignment(horizontal="left", vertical="center")

    # ---- Conditional formatting: highlight matching rows ----
    # The formula is relative so A8 shifts to A9, A10, … for each data row.
    # The AND(LEN($B$5)>0, …) guard prevents every row from highlighting when
    # the search cell is empty (SEARCH("", text) returns 1, i.e. always true).
    #
    # Priority order (first rule = highest priority):
    #   1. Match highlight (yellow)  — overrides banding on matching rows
    #   2. Alternating banding (grey)
    data_range = f"A8:S{last_data_row}"
    match_fill = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[
                "AND(LEN($B$5)>0,"
                "(ISNUMBER(SEARCH($B$5,A8))"
                "+ISNUMBER(SEARCH($B$5,C8))"
                "+ISNUMBER(SEARCH($B$5,L8))"
                "+ISNUMBER(SEARCH($B$5,M8)))>0)"
            ],
            fill=match_fill,
        ),
    )
    banding_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=banding_fill),
    )

    # ---- Formatting ----
    # Same hidden columns as Provisions so the Search view is consistent.
    for col_letter in ("D", "E", "F", "G", "H", "I", "J", "K", "L"):
        ws.column_dimensions[col_letter].hidden = True

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["M"].width = 90
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 14
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["S"].width = 60

    set_column_date_format(ws, ["N", "O", "P"])
    apply_url_hyperlinks(ws, url_col=19)

    ws.auto_filter.ref = f"A7:S{last_data_row}"
    ws.freeze_panes = "A8"


def render_workbook(parsed: ParsedDocument, output_path: Path) -> None:
    wb = Workbook()
    # Use a consistent professional font everywhere — set a workbook default style.
    write_readme(wb, parsed)
    write_metadata(wb, parsed)
    if parsed.document_type == "regulation":
        write_enabling_authorities(wb, parsed)
    write_structure(wb, parsed)
    write_provisions(wb, parsed)
    write_sections(wb, parsed)
    write_definitions(wb, parsed)
    write_amendments(wb, parsed)
    write_schedules(wb, parsed)
    write_search(wb, parsed)

    # Apply Calibri 11 (Excel default) across all cells to keep things crisp.
    base_font = Font(name="Calibri", size=11)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font.bold or cell.font.underline:
                    # Preserve bold headers and hyperlink underline.
                    cell.font = Font(
                        name="Calibri", size=11,
                        bold=cell.font.bold,
                        underline=cell.font.underline,
                        color=cell.font.color,
                    )
                else:
                    cell.font = base_font

    # Apply consistent left + top alignment everywhere. Preserve each cell's
    # existing wrap_text setting (set deliberately on long-text columns and on
    # the README's "Tracking changes" paragraph) and any indent (used on
    # nested Schedule items to show hierarchy), but override horizontal and
    # vertical so headers and ordinary cells all align the same way.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                existing_wrap = bool(cell.alignment and cell.alignment.wrap_text)
                existing_indent = (cell.alignment.indent or 0) if cell.alignment else 0
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=existing_wrap,
                    indent=existing_indent,
                )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ----------------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------------

def prompt_for_input(default_url: str) -> tuple[str, Path | None]:
    """Run the interactive prompts: ask for the source XML and output filename.

    Returns (source, output_override), where ``source`` is the URL or local
    file path to feed to ``load_xml()`` and ``output_override`` is either a
    user-chosen ``Path`` or ``None`` to let the rest of ``main()`` derive a
    default name like ``B-1.01_2026-03-07.xlsx``.

    The function is forgiving in two small ways: it accepts a Justice Laws URL
    without an ``https://`` prefix and silently adds it, and it appends
    ``.xlsx`` to the output filename if the user forgot the extension.
    Ctrl-C or end-of-input at any prompt exits cleanly with status 130 (the
    standard SIGINT exit code).
    """
    print()
    print("=" * 64)
    print("  Justice Canada XML  →  Excel converter")
    print("=" * 64)
    print()
    print("Enter the URL of a Justice Laws XML document, or a local file path.")
    print("Examples:")
    print("  https://laws-lois.justice.gc.ca/eng/XML/B-1.01.xml")
    print("  https://laws-lois.justice.gc.ca/eng/XML/SOR-2021-181.xml")
    print("  ./my_local_copy.xml")
    print(f"  (press Enter to use the default: {default_url})")
    print()

    def ask(prompt: str) -> str:
        try:
            return input(prompt).strip()
        except (EOFError, KeyboardInterrupt):
            print("\nCancelled.")
            sys.exit(130)

    source = ""
    while not source:
        raw = ask("Source URL or file path: ")
        if not raw:
            source = default_url
            print(f"  using default: {source}")
            break
        # Be forgiving when the user pastes a Justice Laws URL without the
        # protocol prefix (a common copy/paste accident).
        if (("laws-lois.justice.gc.ca" in raw)
                and not raw.startswith(("http://", "https://"))):
            raw = "https://" + raw
            print(f"  treating as: {raw}")
        source = raw

    print()
    output_raw = ask(
        "Output filename (press Enter for <Acronym>_YYYY-MM-DD.xlsx): "
    )
    if not output_raw:
        return source, None

    # Helpfully add the extension if the user forgot it.
    if not output_raw.lower().endswith(".xlsx"):
        output_raw += ".xlsx"
    return source, Path(output_raw)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--url", default=None,
                        help="Justice Laws XML URL. If neither --url nor "
                             "--file is given, the script prompts interactively.")
    parser.add_argument("--file", default=None,
                        help="Read XML from a local file instead of downloading.")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: "
                             "<Acronym>_YYYY-MM-DD.xlsx in CWD).")
    parser.add_argument("--lookup-dir", default=None,
                        help="Optional path to a local clone of "
                             "justicecanada/laws-lois-xml/lookup for XRefExternal "
                             "code-to-title resolution.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.lookup_dir:
        load_lookup_from_directory(Path(args.lookup_dir))
        if LOOKUP_TABLE:
            print(f"Loaded {len(LOOKUP_TABLE)} chapter-code mappings from {args.lookup_dir}")

    # Resolve the source (URL or file path) and any output override. If neither
    # --url nor --file was passed on the command line, ask interactively.
    source = args.file or args.url
    output_override: Path | None = Path(args.output) if args.output else None

    if source is None:
        prompted_source, prompted_output = prompt_for_input(default_url=DEFAULT_URL)
        source = prompted_source
        # A command-line --output still wins over the interactive prompt so
        # the two interfaces compose cleanly when used together.
        if output_override is None:
            output_override = prompted_output
        print()

    print(f"Loading XML from {source} ...")
    parsed = parse_document(source)
    print(f"  document type: {parsed.document_type}")

    if output_override is not None:
        output_path = output_override
    else:
        # Default filename: <Acronym>_<today>.xlsx, e.g. CCAA_2026-05-14.xlsx.
        # Fall back through long title -> url code -> document type so we
        # always end up with a sensible name even on oddly-shaped XML.
        name_source = (parsed.short_title or parsed.long_title
                       or parsed.url_code or parsed.document_type)
        stem = to_acronym(name_source) or "Document"
        date_suffix = dt.date.today().isoformat()
        output_path = Path(f"{stem}_{date_suffix}.xlsx")

    render_workbook(parsed, output_path)

    print()
    print(f"Done -> {output_path.resolve()}")
    print(f"  Sections   : {len(parsed.sections):>5}")
    print(f"  Provisions : {len(parsed.provisions):>5}  (all levels combined)")
    print(f"  Definitions: {len(parsed.definitions):>5}")
    print(f"  Amendments : {len(parsed.amendments):>5}")
    print(f"  Headings   : {len(parsed.headings):>5}")
    if parsed.schedules:
        total_items = sum(len(s.items) for s in parsed.schedules)
        print(f"  Schedules  : {len(parsed.schedules):>5}  ({total_items} items total)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
